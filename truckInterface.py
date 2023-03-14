from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import time
import threading
import openpyxl
from PIL import ImageTk, Image
from datetime import datetime
import oracledb

connection = oracledb.connect(
    user='admin',
    password=')Distribucion4',
    dsn='isr4lxwoxgglenv2_low',
    config_dir='opt\OracleCloud\MYDB',
    wallet_location='opt\OracleCloud\MYDB',
    wallet_password=')Distribucion4')

class MainWindow(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.start()
        self.OrdersDone = 0
        self.OrdersPending = 0
        
    def callback(self):
        self.window.quit()  

    def cargaCompleta (self, ZC):
        self.OrdersDone += 1
        cortar = False
        wbprocess = openpyxl.load_workbook("ProcessStatus.xlsx")
        wsQ = wbprocess["VirtualQueue"]
        name = "ZonaCarga"+str(ZC)
        wsprocess = wbprocess[name]
        wsprocess['B1'] = "Libre"
        wsprocess['B2'] = "N/A"
        wsprocess['B3'] = "N/A"
        wsprocess['B4'] = "N/A"
        for rows in wsQ.iter_rows(min_row=2, max_row=40, min_col=3, max_col=3):
            for cell in rows:
                rw = cell.row
                assgLD = wsQ.cell(row=rw, column=4).value
                if (cell.value == "En Zona de Carga") and (assgLD == ZC):
                    orden = wsQ.cell(row=rw, column=1).value
                    wsQ.delete_rows(rw,1)
                    cortar = True
                    break
            if cortar == True:
                break
        wbprocess.save("ProcessStatus.xlsx")
        wbprocess.close()
        wbordenes = openpyxl.load_workbook("DB_Clientes_NoOrden.xlsx")
        wsordenes = wbordenes["Ordenes"]
        for rowsOr in wsordenes.iter_rows(min_row=2, max_row=40, min_col=3, max_col=3):
            for cellOr in rowsOr:
                rwOr = cellOr.row
                if cellOr.value == orden:
                    wsordenes.cell(row=rwOr, column=4, value="Completada")
                    wsordenes.cell(row=rwOr, column=5, value=None)
        wbordenes.save("DB_Clientes_NoOrden.xlsx")
        wbordenes.close()




    def warning(self):
        warnWin = messagebox.askokcancel(title="Program Stopped", message="The STOP button was pressed. Do you want to quit the program?", icon='warning')
        if warnWin:
            self.window.quit()

    def run(self):
        self.window = Tk()
        self.window.protocol("WM_DELETE_WINDOW", self.callback)
        width = self.window.winfo_screenwidth()
        height = self.window.winfo_screenheight()
        self.window.geometry("%dx%d" % (width, height))
        self.window.title("Truck GUI")
        self.window.configure(bg='#242529')
        self.window.iconbitmap('TruckFront.ico')
        
        self.frame1 = Frame(self.window)
        self.frame1.config(bg='#242529')
        self.frame1.pack(fill=X)
        self.frame2 = Frame(self.window)
        self.frame2.config(bg='#242529')
        self.frame2.pack(fill=X, padx = 70, pady = 20)
        self.frame3 = Frame(self.window)
        self.frame3.config(bg='#242529')
        self.frame3.pack(fill=X, padx = 100)
        self.frame4 = Frame(self.window)
        self.frame4.config(bg='#242529')
        self.frame4.pack(fill=X)

        self.frame1.columnconfigure(0, weight=1)
        self.frame1.columnconfigure(1, weight=1)
        self.frame1.columnconfigure(2, weight=1)

        self.frame2.rowconfigure(0, weight = 1)
        self.frame2.rowconfigure(1, weight = 2)
        self.frame2.rowconfigure(2, weight = 1)
        self.frame2.rowconfigure(3, weight = 1)
        self.frame2.rowconfigure(4, weight = 1)
        self.frame2.rowconfigure(5, weight = 1)
        self.frame2.rowconfigure(6, weight = 1)
        self.frame2.columnconfigure(0, weight = 1)
        self.frame2.columnconfigure(1, weight = 1)
        self.frame2.columnconfigure(2, weight = 1)
        self.frame2.columnconfigure(3, weight = 1)
        self.frame2.columnconfigure(4, weight = 1)
        self.frame2.columnconfigure(5, weight = 1)

        self.frame3.rowconfigure(0, weight = 1)
        self.frame3.rowconfigure(1, weight = 2)
        self.frame3.rowconfigure(2, weight = 1)
        self.frame3.rowconfigure(3, weight = 1)
        self.frame3.rowconfigure(4, weight = 1)
        self.frame3.rowconfigure(5, weight = 1)
        self.frame3.rowconfigure(6, weight = 1)

        self.frame3.columnconfigure(0, weight = 1)
        self.frame3.columnconfigure(1, weight = 1)
        self.frame3.columnconfigure(2, weight = 1)
        self.frame3.columnconfigure(3, weight = 1)
        self.frame3.columnconfigure(4, weight = 1)

        self.frame4.columnconfigure(0, weight = 1)
        self.frame4.columnconfigure(1, weight = 1)
        self.frame4.columnconfigure(2, weight = 1)

        Titulo = Label(self.frame1,text="CONTROL DE CARGA DE TRANSPORTE", font=("Verdana", 25, 'bold'), bg='#242529', fg='white')
        Titulo.grid(row=0, column=1, pady=15)

        Extra = Label(self.frame1, text="                    ", font=("Verdana", 20, 'bold'), bg='#242529', fg='white')
        Extra.grid(row=0, column=0)

        self.TxtTime = StringVar()
        self.TxtTime.set(datetime.now().strftime("%d/%m/%Y %H:%M"))
        self.TimeLbl = Label(self.frame1, textvariable=self.TxtTime, font=("Verdana", 13), bg='#242529', fg='white')
        self.TimeLbl.grid(row=0, column=2, sticky=E, padx=20)

        STLD1 = Label(self.frame2, text="Zona de Carga 1", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD2 = Label(self.frame2, text="Zona de Carga 2", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD3 = Label(self.frame2, text="Zona de Carga 3", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD4 = Label(self.frame2, text="Zona de Carga 4", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD5 = Label(self.frame2, text="Zona de Carga 5", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD6 = Label(self.frame2, text="Zona de Carga 6", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STLD1.grid(row=0, column=0)
        STLD2.grid(row=0, column=1)
        STLD3.grid(row=0, column=2)
        STLD4.grid(row=0, column=3)
        STLD5.grid(row=0, column=4)
        STLD6.grid(row=0, column=5)
        sep1_1 = ttk.Separator(self.frame2, orient='horizontal')
        sep1_2 = ttk.Separator(self.frame2, orient='horizontal')
        sep1_3 = ttk.Separator(self.frame2, orient='vertical')
        sep1_4 = ttk.Separator(self.frame2, orient='vertical')
        sep1_5 = ttk.Separator(self.frame2, orient='vertical')
        sep1_6 = ttk.Separator(self.frame2, orient='vertical')
        sep1_7 = ttk.Separator(self.frame2, orient='vertical')
        sep1_8 = ttk.Separator(self.frame2, orient='vertical')
        sep1_9 = ttk.Separator(self.frame2, orient='vertical')
        sep1_1.grid(column=0, row=0, columnspan=6, sticky=N+E+W)
        sep1_2.grid(column=0, row=6, columnspan=6, sticky=N+E+W)
        sep1_3.grid(row = 0, column = 0, rowspan = 6, sticky=N+S+W)
        sep1_4.grid(row = 0, column = 1, rowspan = 6, sticky=N+S+W)
        sep1_5.grid(row = 0, column = 2, rowspan = 6, sticky=N+S+W)
        sep1_6.grid(row = 0, column = 3, rowspan = 6, sticky=N+S+W)
        sep1_7.grid(row = 0, column = 4, rowspan = 6, sticky=N+S+W)
        sep1_8.grid(row = 0, column = 5, rowspan = 6, sticky=N+S+W)
        sep1_9.grid(row = 0, column = 6, rowspan = 6, sticky=N+S+W)

        STWZ1 = Label(self.frame3, text="Zona de Espera 1", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STWZ2 = Label(self.frame3, text="Zona de Espera 2", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STWZ3 = Label(self.frame3, text="Zona de Espera 3", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STWZ4 = Label(self.frame3, text="Zona de Espera 4", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STWZ5 = Label(self.frame3, text="Zona de Espera 5", font=("Verdana",13, 'bold'), bg='#242529', fg='white')
        STWZ1.grid(row=0, column=0)
        STWZ2.grid(row=0, column=1)
        STWZ3.grid(row=0, column=2)
        STWZ4.grid(row=0, column=3)
        STWZ5.grid(row=0, column=4)
        sep2_1 = ttk.Separator(self.frame3, orient='horizontal')
        sep2_2 = ttk.Separator(self.frame3, orient='horizontal')
        sep2_3 = ttk.Separator(self.frame3, orient='vertical')
        sep2_4 = ttk.Separator(self.frame3, orient='vertical')
        sep2_5 = ttk.Separator(self.frame3, orient='vertical')
        sep2_6 = ttk.Separator(self.frame3, orient='vertical')
        sep2_7 = ttk.Separator(self.frame3, orient='vertical')
        sep2_8 = ttk.Separator(self.frame3, orient='vertical')
        sep2_1.grid(column=0, row=0, columnspan=6, sticky=N+E+W)
        sep2_2.grid(column=0, row=7, columnspan=6, sticky=N+E+W)
        sep2_3.grid(row = 0, column = 0, rowspan = 7, sticky=N+S+W)
        sep2_4.grid(row = 0, column = 1, rowspan = 7, sticky=N+S+W)
        sep2_5.grid(row = 0, column = 2, rowspan = 7, sticky=N+S+W)
        sep2_6.grid(row = 0, column = 3, rowspan = 7, sticky=N+S+W)
        sep2_7.grid(row = 0, column = 4, rowspan = 7, sticky=N+S+W)
        sep2_8.grid(row = 0, column = 5, rowspan = 7, sticky=N+S+W)

        self.LZFree = [0 for i in range(6)]
        for i in range(6):
            self.LZFree[i] = Button(self.frame2, text="Carga Completa", font=("Verdana",13, 'bold'), borderwidth=4, bg='#f9e000', state='disabled', command=lambda: self.cargaCompleta(i+1))
            self.LZFree[i].grid(column=i, row=6, pady=10)
        
        self.imgtruck = ImageTk.PhotoImage(Image.open("truck.png").resize((148,68)))
        self.imgempty = ImageTk.PhotoImage(Image.open("point.png").resize((148,68)))
        self.truckLD = [0 for i in range(6)]
        for i in range(6):
            self.truckLD[i] = Label(self.frame2, image=self.imgempty, bg='#242529')
            self.truckLD[i].image = self.imgempty
            self.truckLD[i].grid(column=i, row=1)


        self.TextEstadoLD = [StringVar() for i in range(6)]
        for i in range(6):
            self.TextEstadoLD[i].set("Estado: En Espera")
            
        self.estadoLD1 = Label(self.frame2, textvariable=self.TextEstadoLD[0], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD2 = Label(self.frame2, textvariable=self.TextEstadoLD[1], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD3 = Label(self.frame2, textvariable=self.TextEstadoLD[2], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD4 = Label(self.frame2, textvariable=self.TextEstadoLD[3], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD5 = Label(self.frame2, textvariable=self.TextEstadoLD[4], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD6 = Label(self.frame2, textvariable=self.TextEstadoLD[5], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoLD1.grid(column=0, row=2, sticky=W, padx=5)
        self.estadoLD2.grid(column=1, row=2, sticky=W, padx=5)
        self.estadoLD3.grid(column=2, row=2, sticky=W, padx=5)
        self.estadoLD4.grid(column=3, row=2, sticky=W, padx=5)
        self.estadoLD5.grid(column=4, row=2, sticky=W, padx=5)
        self.estadoLD6.grid(column=5, row=2, sticky=W, padx=5)


        self.TextOrdenLD = [StringVar() for i in range(6)]
        for i in range(6):
            self.TextOrdenLD[i].set("Orden: N/A")

        self.ordenLD1 = Label(self.frame2, textvariable=self.TextOrdenLD[0], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD2 = Label(self.frame2, textvariable=self.TextOrdenLD[1], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD4 = Label(self.frame2, textvariable=self.TextOrdenLD[2], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD3 = Label(self.frame2, textvariable=self.TextOrdenLD[3], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD5 = Label(self.frame2, textvariable=self.TextOrdenLD[4], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD6 = Label(self.frame2, textvariable=self.TextOrdenLD[5], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenLD1.grid(column=0, row=3, sticky=W, padx=5)
        self.ordenLD2.grid(column=1, row=3, sticky=W, padx=5)
        self.ordenLD3.grid(column=2, row=3, sticky=W, padx=5)
        self.ordenLD4.grid(column=3, row=3, sticky=W, padx=5)
        self.ordenLD5.grid(column=4, row=3, sticky=W, padx=5)
        self.ordenLD6.grid(column=5, row=3, sticky=W, padx=5)


        self.botOrdenLD = [Button(self.frame2, text='*', font=("Verdana", 10, 'bold'), state='disabled') for i in range(6)]
        for i in range(6):
            self.botOrdenLD[i].grid(column=i, row=3, sticky=E, padx=5, pady=5)


        self.TextPlacasLD = [StringVar() for i in range(6)]
        for i in range(6):
            self.TextPlacasLD[i].set("Placas: N/A")

        self.placasLD1 = Label(self.frame2, textvariable=self.TextPlacasLD[0], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD2 = Label(self.frame2, textvariable=self.TextPlacasLD[1], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD3 = Label(self.frame2, textvariable=self.TextPlacasLD[2], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD4 = Label(self.frame2, textvariable=self.TextPlacasLD[3], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD5 = Label(self.frame2, textvariable=self.TextPlacasLD[4], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD6 = Label(self.frame2, textvariable=self.TextPlacasLD[5], font=("Verdana",13), bg='#242529', fg='white')
        self.placasLD1.grid(column=0, row=4, sticky=W, padx=5)
        self.placasLD2.grid(column=1, row=4, sticky=W, padx=5)
        self.placasLD3.grid(column=2, row=4, sticky=W, padx=5)
        self.placasLD4.grid(column=3, row=4, sticky=W, padx=5)
        self.placasLD5.grid(column=4, row=4, sticky=W, padx=5)
        self.placasLD6.grid(column=5, row=4, sticky=W, padx=5)


        # self.TextTiempoLD = [StringVar() for i in range(6)]
        # for i in range(6):
        #     self.TextTiempoLD[i].set("Tiempo Restante: N/A")

        # self.tiempoLD1 = Label(self.frame2, textvariable=self.TextTiempoLD[0], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD2 = Label(self.frame2, textvariable=self.TextTiempoLD[1], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD3 = Label(self.frame2, textvariable=self.TextTiempoLD[2], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD4 = Label(self.frame2, textvariable=self.TextTiempoLD[3], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD5 = Label(self.frame2, textvariable=self.TextTiempoLD[4], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD6 = Label(self.frame2, textvariable=self.TextTiempoLD[5], font=("Verdana",13), bg='#242529', fg='white')
        # self.tiempoLD1.grid(column=0, row=5, sticky=W, padx=5)
        # self.tiempoLD2.grid(column=1, row=5, sticky=W, padx=5)
        # self.tiempoLD3.grid(column=2, row=5, sticky=W, padx=5)
        # self.tiempoLD4.grid(column=3, row=5, sticky=W, padx=5)
        # self.tiempoLD5.grid(column=4, row=5, sticky=W, padx=5)
        # self.tiempoLD6.grid(column=5, row=5, sticky=W, padx=5)

        self.truck1WZ = Label(self.frame3, image=self.imgempty, bg='#242529')
        self.truck2WZ = Label(self.frame3, image=self.imgempty, bg='#242529')
        self.truck3WZ = Label(self.frame3, image=self.imgempty, bg='#242529')
        self.truck4WZ = Label(self.frame3, image=self.imgempty, bg='#242529')
        self.truck5WZ = Label(self.frame3, image=self.imgempty, bg='#242529')
        self.truck1WZ.image = self.imgempty
        self.truck2WZ.image = self.imgempty
        self.truck3WZ.image = self.imgempty
        self.truck4WZ.image = self.imgempty
        self.truck5WZ.image = self.imgempty
        self.truck1WZ.grid(column=0, row=1)
        self.truck2WZ.grid(column=1, row=1)
        self.truck3WZ.grid(column=2, row=1)
        self.truck4WZ.grid(column=3, row=1)
        self.truck5WZ.grid(column=4, row=1)


        self.TextEstadoWZ = [StringVar() for i in range(5)]
        for i in range(5):
            self.TextEstadoWZ[i].set("Estado: Vacio")

        self.estadoWZ1 = Label(self.frame3, textvariable=self.TextEstadoWZ[0], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoWZ2 = Label(self.frame3, textvariable=self.TextEstadoWZ[1], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoWZ3 = Label(self.frame3, textvariable=self.TextEstadoWZ[2], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoWZ4 = Label(self.frame3, textvariable=self.TextEstadoWZ[3], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoWZ5 = Label(self.frame3, textvariable=self.TextEstadoWZ[4], font=("Verdana",13), bg='#242529', fg='white')
        self.estadoWZ1.grid(column=0, row=2, sticky=W, padx=5)
        self.estadoWZ2.grid(column=1, row=2, sticky=W, padx=5)
        self.estadoWZ3.grid(column=2, row=2, sticky=W, padx=5)
        self.estadoWZ4.grid(column=3, row=2, sticky=W, padx=5)
        self.estadoWZ5.grid(column=4, row=2, sticky=W, padx=5)


        self.TextOrdenWZ = [StringVar() for i in range(5)]
        for i in range(5):
            self.TextOrdenWZ[i].set("Orden: N/A")

        self.ordenWZ1 = Label(self.frame3, textvariable=self.TextOrdenWZ[0], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenWZ2 = Label(self.frame3, textvariable=self.TextOrdenWZ[1], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenWZ3 = Label(self.frame3, textvariable=self.TextOrdenWZ[2], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenWZ4 = Label(self.frame3, textvariable=self.TextOrdenWZ[3], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenWZ5 = Label(self.frame3, textvariable=self.TextOrdenWZ[4], font=("Verdana",13), bg='#242529', fg='white')
        self.ordenWZ1.grid(column=0, row=3, sticky=W, padx=5)
        self.ordenWZ2.grid(column=1, row=3, sticky=W, padx=5)
        self.ordenWZ3.grid(column=2, row=3, sticky=W, padx=5)
        self.ordenWZ4.grid(column=3, row=3, sticky=W, padx=5)
        self.ordenWZ5.grid(column=4, row=3, sticky=W, padx=5)


        self.TextPlacasWZ = [StringVar() for i in range(5)]
        for i in range(5):
            self.TextPlacasWZ[i].set("Placas: N/A")

        self.placasWZ1 = Label(self.frame3, textvariable=self.TextPlacasWZ[0], font=("Verdana",13), bg='#242529', fg='white')
        self.placasWZ2 = Label(self.frame3, textvariable=self.TextPlacasWZ[1], font=("Verdana",13), bg='#242529', fg='white')
        self.placasWZ3 = Label(self.frame3, textvariable=self.TextPlacasWZ[2], font=("Verdana",13), bg='#242529', fg='white')
        self.placasWZ4 = Label(self.frame3, textvariable=self.TextPlacasWZ[3], font=("Verdana",13), bg='#242529', fg='white')
        self.placasWZ5 = Label(self.frame3, textvariable=self.TextPlacasWZ[4], font=("Verdana",13), bg='#242529', fg='white')
        self.placasWZ1.grid(column=0, row=4, sticky=W, padx=5)
        self.placasWZ2.grid(column=1, row=4, sticky=W, padx=5)
        self.placasWZ3.grid(column=2, row=4, sticky=W, padx=5)
        self.placasWZ4.grid(column=3, row=4, sticky=W, padx=5)
        self.placasWZ5.grid(column=4, row=4, sticky=W, padx=5)


        self.TextSiguienteWZ = [StringVar() for i in range(5)]
        for i in range(5):
            self.TextSiguienteWZ[i].set("A Zona de Carga: N/A")

        self.siguienteWZ1 = Label(self.frame3, textvariable=self.TextSiguienteWZ[0], font=("Verdana",13), bg='#242529', fg='white')
        self.siguienteWZ2 = Label(self.frame3, textvariable=self.TextSiguienteWZ[1], font=("Verdana",13), bg='#242529', fg='white')
        self.siguienteWZ3 = Label(self.frame3, textvariable=self.TextSiguienteWZ[2], font=("Verdana",13), bg='#242529', fg='white')
        self.siguienteWZ4 = Label(self.frame3, textvariable=self.TextSiguienteWZ[3], font=("Verdana",13), bg='#242529', fg='white')
        self.siguienteWZ5 = Label(self.frame3, textvariable=self.TextSiguienteWZ[4], font=("Verdana",13), bg='#242529', fg='white')
        self.siguienteWZ1.grid(column=0, row=5, sticky=W, padx=5)
        self.siguienteWZ2.grid(column=1, row=5, sticky=W, padx=5)
        self.siguienteWZ3.grid(column=2, row=5, sticky=W, padx=5)
        self.siguienteWZ4.grid(column=3, row=5, sticky=W, padx=5)
        self.siguienteWZ5.grid(column=4, row=5, sticky=W, padx=5)


        imgstop = ImageTk.PhotoImage(Image.open("StopImg.png").resize((130,130)))
        botonStop = Button(self.frame4, image=imgstop, bg='#242529', borderwidth=0, command=self.warning)
        botonStop.grid(column=0, pady=30, padx=80, sticky=W)

        self.TextOrdsPen = StringVar()
        self.TextOrdsPen.set("Ordenes Pendientes: 0")
        self.OrdsPen = Label(self.frame4, textvariable=self.TextOrdsPen, font=("Verdana",15), bg='#242529', fg='white')
        self.OrdsPen.grid(row=0, column=1)
        self.TextOrdsDesp = StringVar()
        self.TextOrdsDesp.set("Ordenes Despachadas: 0")
        self.OrdsDesp = Label(self.frame4, textvariable=self.TextOrdsDesp, font=("Verdana",15), bg='#242529', fg='white')
        self.OrdsDesp.grid(row=0, column=2)

        self.window.mainloop()


    def update(self):
        now = datetime.now()
        mytime = now.strftime("%d/%m/%Y %H:%M")
        self.TxtTime.set(mytime)

        i = 0
        with connection.cursor() as cursor:
            for row in cursor.execute('select estado, orden_en_proceso, placas from zona_de_carga'):
                if row[0] == 'Cargando':
                    self.truckLD[i].configure(image=self.imgtruck)
                    self.LZFree[i]["state"] = 'normal'
                else:
                    self.truckLD[i].configure(image=self.imgempty)
                    self.LZFree[i]["state"] = 'disabled'
                self.TextEstadoLD[i].set("Estado: " + row[0])
                self.TextOrdenLD[i].set("Orden: " + str(row[1]))
                self.TextPlacasLD[i].set("Placas: " + str(row[2]))
                if row[1] != "null":
                    self.botOrdenLD[i]["state"] = 'normal'
                else:
                    self.botOrdenLD[i]["state"] = 'disabled'
                # print(i)
                # print(row[0])
                # print(row[1])
                i+=1
                
        i = 0
        with connection.cursor() as cursor:
            for row in cursor.execute('select zona_espera, estado, orden, placas, zona_carga from zona_de_espera'):
                if row[1] == "En Espera":
                    self.truck1WZ.configure(image=self.imgtruck)
                else:
                    self.truck1WZ.configure(image=self.imgempty)
        
                self.TextEstadoWZ[i].set("Estado: " + row[1])
                self.TextOrdenWZ[i].set("Orden: " + str(row[2]))
                self.TextPlacasWZ[i].set("Placas: " + str(row[3]))
                self.TextSiguienteWZ[i].set("A Zona de Carga: "+ str(row[4]))
                i+=1
      
        self.OrdersPending = -1
        with connection.cursor() as cursor:
            for row in cursor.execute('select estatus_orden from ordenes'):
                if (row[0]!="Completada"):
                    self.OrdersPending += 1
        self.TextOrdsPen.set("Ordenes Pendientes: "+str(self.OrdersPending))
        self.TextOrdsDesp.set("Ordenes Despachadas: "+str(self.OrdersDone))

MW = MainWindow()
time.sleep(1)
while True:
    try:
        if MW.window.winfo_exists():
            MW.update()
    except Exception as e: 
        print(e)
        MW.callback()
        exit()